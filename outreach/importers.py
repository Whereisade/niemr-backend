from __future__ import annotations

import csv
import io

def read_tabular_file(uploaded_file):
    """Read CSV or XLSX into list[dict].

    - CSV: utf-8
    - XLSX: requires openpyxl
    """
    if not uploaded_file:
        raise ValueError("file is required")

    filename = (uploaded_file.name or "").lower()
    if filename.endswith(".csv"):
        buf = io.StringIO(uploaded_file.read().decode("utf-8"))
        reader = csv.DictReader(buf)
        return list(reader)
    if filename.endswith(".xlsx"):
        import openpyxl  # already in requirements

        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i].strip().lower(): values[i] for i in range(len(headers))})
        return rows
    if filename.endswith(".xls"):
        raise ValueError("Legacy .xls is not supported. Please save as .xlsx or CSV.")
    raise ValueError("Unsupported file format. Please upload CSV or Excel (.xlsx) file.")
