from django.utils.dateparse import parse_datetime
from django.utils import timezone
from django.db.models import Q, Case, When, IntegerField, Value, Subquery
from rest_framework import viewsets, mixins, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.contrib.auth import get_user_model
from billing.models import Service, Price, Charge
from accounts.enums import UserRole
from patients.models import SystemHMO, HMOTier
from facilities.permissions_utils import has_facility_permission
from billing.services.pricing import get_service_price_info, resolve_price
from .models import Appointment
from .serializers import (
    AppointmentSerializer,
    AppointmentUpdateSerializer,
    AppointmentListSerializer,
)
from .permissions import IsStaff, CanViewAppointment
from .enums import ApptStatus
from decimal import Decimal
from billing.models import Service
from .services.notify import (
    send_confirmation,
    send_reminder,
    send_provider_assignment,
    send_completed,
    send_cancelled,
    send_no_show,
)
from notifications.services.notify import notify_user, notify_users, notify_facility_roles, notify_patient
from notifications.enums import Topic, Priority
from accounts.enums import UserRole

User = get_user_model()


class AppointmentViewSet(
    viewsets.GenericViewSet,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin,
):
    queryset = Appointment.objects.select_related(
        "patient", "facility", "provider", "created_by"
    )
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == "list":
            return AppointmentListSerializer
        if self.action in ("update", "partial_update"):
            return AppointmentUpdateSerializer
        return AppointmentSerializer


    def ensure_appointment_services():
        """
        Create all 18 appointment service types if they don't exist.
        Call this once during deployment or add to migration.
        """
        SERVICE_TYPES = {
            "APPT:CONSULT_STD": ("Standard Consultation", 5000),
            "APPT:CONSULT_FOLLOW_UP": ("Follow-up Consultation", 3000),
            "APPT:CONSULT_EMERGENCY": ("Emergency Consultation", 10000),
            "APPT:CONSULT_SPECIALIST": ("Specialist Consultation", 8000),
            "APPT:CONSULT_PEDIATRIC": ("Pediatric Consultation", 6000),
            "APPT:ANNUAL_CHECKUP": ("Annual Health Checkup", 15000),
            "APPT:PHYSICAL_EXAM": ("Physical Examination", 5000),
            "APPT:WELLNESS_VISIT": ("Wellness Visit", 4000),
            "APPT:IMMUNIZATION": ("Immunization/Vaccination", 3000),
            "APPT:LAB_COLLECTION": ("Lab Sample Collection", 2000),
            "APPT:X_RAY_SCREENING": ("X-Ray Screening", 5000),
            "APPT:DENTAL_CHECKUP": ("Dental Checkup", 4000),
            "APPT:VISION_SCREENING": ("Vision Screening", 3000),
            "APPT:HEARING_TEST": ("Hearing Test", 3000),
            "APPT:COUNSELING_SESSION": ("Counseling Session", 6000),
            "APPT:NUTRITION_CONSULT": ("Nutrition Consultation", 5000),
            "APPT:THERAPY_SESSION": ("Therapy Session (Physical/Occupational)", 7000),
            "APPT:ADMIN_HMO_REVIEW": ("Administrative/HMO Review", 1000),
        }
        
        created_count = 0
        for code, (name, price) in SERVICE_TYPES.items():
            service, created = Service.objects.get_or_create(
                code=code,
                defaults={
                    "name": name,
                    "default_price": Decimal(str(price)),
                    "is_active": True
                }
            )
            if created:
                created_count += 1
        
        return created_count

    def get_queryset(self):
        q = self.queryset
        u = self.request.user

        # Role-based filtering
        if u.role == "PATIENT":
            base_patient = getattr(u, "patient_profile", None)
            if base_patient:
                # Patient can see own appointments and dependents'
                q = q.filter(
                    Q(patient=base_patient) | Q(patient__parent_patient=base_patient)
                )
            else:
                q = q.none()
        elif u.facility_id:
            # Facility-scoped users
            q = q.filter(facility_id=u.facility_id)

            # IMPORTANT: Facility doctors should only see appointments assigned to them.
            # Other facility roles (SUPER_ADMIN, ADMIN, NURSE, FRONTDESK, etc.) can see all.
            if u.role == "DOCTOR":
                # Primary: appointment.provider points to the assigned doctor
                # Fallback: if older data only has encounter.provider, include those too.
                try:
                    from encounters.models import Encounter

                    enc_ids = Encounter.objects.filter(provider_id=u.id).values("id")
                    q = q.filter(Q(provider_id=u.id) | Q(encounter_id__in=Subquery(enc_ids)))
                except Exception:
                    q = q.filter(provider_id=u.id)

            # Optional: facility staff can still request only their own appointments via ?mine=true
            # (ignored for doctors because they are always scoped to their own).
            elif self.request.query_params.get("mine") in ("true", "True", "1"):
                q = q.filter(provider_id=u.id)
        else:
            # Independent provider without facility - only own appointments
            q = q.filter(provider_id=u.id)

        # Query params filtering
        patient_id = self.request.query_params.get("patient")
        provider_id = self.request.query_params.get("provider")
        status_ = self.request.query_params.get("status")
        start = self.request.query_params.get("start")
        end = self.request.query_params.get("end")
        s = self.request.query_params.get("s") or self.request.query_params.get("q")
        date_filter = self.request.query_params.get("date")

        if patient_id:
            q = q.filter(patient_id=patient_id)
        if provider_id:
            q = q.filter(provider_id=provider_id)
        if status_:
            # Handle both upper and lower case status values
            q = q.filter(status__iexact=status_)
        if start:
            parsed_start = parse_datetime(start)
            q = q.filter(start_at__gte=parsed_start or start)
        if end:
            parsed_end = parse_datetime(end)
            q = q.filter(end_at__lte=parsed_end or end)
        if s:
            q = q.filter(
                Q(reason__icontains=s)
                | Q(notes__icontains=s)
                | Q(patient__first_name__icontains=s)
                | Q(patient__last_name__icontains=s)
            )

        # Date presets for convenience
        if date_filter:
            today = timezone.now().date()
            if date_filter == "today":
                q = q.filter(start_at__date=today)
            elif date_filter == "tomorrow":
                q = q.filter(start_at__date=today + timezone.timedelta(days=1))
            elif date_filter == "this_week":
                week_start = today - timezone.timedelta(days=today.weekday())
                week_end = week_start + timezone.timedelta(days=6)
                q = q.filter(start_at__date__gte=week_start, start_at__date__lte=week_end)
            elif date_filter == "next_7d":
                q = q.filter(
                    start_at__date__gte=today,
                    start_at__date__lte=today + timezone.timedelta(days=7),
                )
            # 'all' or unknown values = no date filter

        # Order list with active (new/current) appointments first
        status_rank = Case(
            When(status__iexact=ApptStatus.CHECKED_IN, then=Value(0)),
            When(status__iexact=ApptStatus.SCHEDULED, then=Value(1)),
            When(status__iexact=ApptStatus.COMPLETED, then=Value(2)),
            When(status__iexact=ApptStatus.CANCELLED, then=Value(3)),
            When(status__iexact=ApptStatus.NO_SHOW, then=Value(4)),
            default=Value(99),
            output_field=IntegerField(),
        )

        return q.annotate(_status_rank=status_rank).order_by("_status_rank", "start_at", "id")


    def _attach_prefetched_encounters(self, appts):
        """Attach Encounter objects to each appointment (as _prefetched_encounter) to avoid N+1."""
        try:
            from encounters.models import Encounter
        except Exception:
            return

        encounter_ids = [a.encounter_id for a in appts if getattr(a, "encounter_id", None)]
        if not encounter_ids:
            return

        enc_map = {
            e.id: e
            for e in Encounter.objects.filter(id__in=encounter_ids).select_related("nurse", "provider")
        }

        for a in appts:
            if getattr(a, "encounter_id", None):
                a._prefetched_encounter = enc_map.get(a.encounter_id)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            # page is already a list
            self._attach_prefetched_encounters(page)
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        objs = list(queryset)
        self._attach_prefetched_encounters(objs)
        serializer = self.get_serializer(objs, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """
        Create appointment with proper patient/dependent handling.
        """
        user = request.user
        data = request.data.copy()

        if user.role == "PATIENT":
            base_patient = getattr(user, "patient_profile", None)
            if not base_patient:
                return Response(
                    {"detail": "No patient profile linked to this user."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            raw_patient_id = data.get("patient")
            target_patient_id = None
            if raw_patient_id not in (None, "", "null"):
                try:
                    target_patient_id = int(raw_patient_id)
                except (TypeError, ValueError):
                    return Response(
                        {"detail": "Invalid patient id."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            if target_patient_id is None or target_patient_id == base_patient.id:
                data["patient"] = base_patient.id
            else:
                allowed_ids = set(base_patient.dependents.values_list("id", flat=True))
                if target_patient_id not in allowed_ids:
                    return Response(
                        {
                            "detail": (
                                "You can only book appointments for yourself "
                                "or your registered dependents."
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                data["patient"] = target_patient_id
        else:
            self.permission_classes = [IsAuthenticated, IsStaff]
            self.check_permissions(request)

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        resp = Response(
            serializer.data,
            status=status.HTTP_201_CREATED,
            headers=self.get_success_headers(serializer.data),
        )

        # Auto-link patient to facility
        try:
            appt = Appointment.objects.select_related("patient", "facility").get(
                id=resp.data["id"]
            )
            patient = appt.patient
            if appt.facility_id and (
                not patient.facility_id or patient.facility_id != appt.facility_id
            ):
                patient.facility = appt.facility
                patient.save(update_fields=["facility"])
        except Exception:
            pass

        # Send confirmation email
        try:
            appt = Appointment.objects.get(id=resp.data["id"])
            send_confirmation(appt)
            # Provider assignment email (doctor/nurse/etc)
            send_provider_assignment(appt)
        except Exception:
            pass
        # In-app notifications (facility staff + provider + patient)
        try:
            appt = Appointment.objects.select_related(
                "patient", "facility", "provider", "patient__user"
            ).get(id=resp.data["id"])

            facility_id = appt.facility_id
            patient_name = getattr(appt.patient, "full_name", None) or f"Patient #{appt.patient_id}"
            when = appt.start_at.strftime("%Y-%m-%d %H:%M") if appt.start_at else ""

            title = "New appointment scheduled"
            body = f"{patient_name} • {when}\nReason: {appt.reason or '-'}"
            data_payload = {"appointment_id": appt.id, "patient_id": appt.patient_id}
            group_key = f"APPT:{appt.id}:CREATED"

            # Provider
            if appt.provider_id:
                notify_user(
                    user=appt.provider,
                    topic=Topic.APPOINTMENT_CONFIRMED,
                    priority=Priority.NORMAL,
                    title=title,
                    body=body,
                    facility_id=facility_id,
                    data=data_payload,
                    action_url="/facility/appointments",
                    group_key=group_key,
                )

            # Facility staff
            if facility_id:
                staff_roles = [
                    UserRole.SUPER_ADMIN,
                    UserRole.ADMIN,
                    UserRole.FRONTDESK,
                    UserRole.NURSE,
                ]
                staff_users = (
                    User.objects.filter(facility_id=facility_id, role__in=staff_roles)
                    .exclude(id=appt.provider_id)
                    .distinct()
                )
                notify_users(
                    users=staff_users,
                    topic=Topic.APPOINTMENT_CONFIRMED,
                    priority=Priority.NORMAL,
                    title=title,
                    body=body,
                    facility_id=facility_id,
                    data=data_payload,
                    action_url="/facility/appointments",
                    group_key=group_key,
                )

            # Patient (and guardian, if dependent)
            if appt.patient:
                notify_patient(
                    patient=appt.patient,
                    topic=Topic.APPOINTMENT_CONFIRMED,
                    priority=Priority.LOW,
                    title="Appointment booked",
                    body=f"Your appointment is scheduled for {when}.",
                    facility_id=facility_id,
                    data=data_payload,
                    action_url="/patient/appointments",
                    group_key=group_key,
                )
        except Exception:
            pass

        return resp

    def retrieve(self, request, *args, **kwargs):
        obj = self.get_object()
        self.permission_classes = [IsAuthenticated, CanViewAppointment]
        self.check_object_permissions(request, obj)
        return Response(AppointmentSerializer(obj, context={"request": request}).data)

    def update(self, request, *args, **kwargs):
        appt = self.get_object()

        # Only staff can update appointments (patients book via create and can cancel).
        self.permission_classes = [IsAuthenticated, IsStaff]
        self.check_permissions(request)

        old_start = appt.start_at
        old_end = appt.end_at
        old_provider_id = appt.provider_id

        resp = super().update(request, *args, **kwargs)

        try:
            appt.refresh_from_db()
            changed_time = (old_start != appt.start_at) or (old_end != appt.end_at)
            changed_provider = (old_provider_id != appt.provider_id)

            if changed_time or changed_provider:
                new_when = appt.start_at.strftime("%Y-%m-%d %H:%M") if appt.start_at else ""
                old_when = old_start.strftime("%Y-%m-%d %H:%M") if old_start else ""
                payload = {"appointment_id": appt.id, "patient_id": appt.patient_id}
                group_key = f"APPT:{appt.id}:UPDATED"

                if changed_time and appt.patient:
                    body = (
                        f"Your appointment has been rescheduled to {new_when}."
                        if not old_when
                        else f"Your appointment has been rescheduled from {old_when} to {new_when}."
                    )
                    notify_patient(
                        patient=appt.patient,
                        topic=Topic.APPOINTMENT_RESCHEDULED,
                        priority=Priority.NORMAL,
                        title="Appointment rescheduled",
                        body=body,
                        facility_id=appt.facility_id,
                        data=payload,
                        action_url="/patient/appointments",
                        group_key=group_key,
                    )

                # Provider changes (if any)
                if changed_provider and appt.provider_id:
                    notify_user(
                        user=appt.provider,
                        topic=Topic.STAFF_ASSIGNED,
                        priority=Priority.NORMAL,
                        title="New appointment assigned",
                        body=f"You have been assigned an appointment for {new_when}.",
                        facility_id=appt.facility_id,
                        data=payload,
                        action_url="/facility/appointments",
                        group_key=group_key,
                    )
                    # Provider email (best-effort)
                    try:
                        send_provider_assignment(appt)
                    except Exception:
                        pass

                # Ops feed
                if appt.facility_id:
                    patient_name = " ".join([p for p in [getattr(appt.patient, 'first_name', ''), getattr(appt.patient, 'middle_name', ''), getattr(appt.patient, 'last_name', '')] if p]).strip()
                    notify_facility_roles(
                        facility_id=appt.facility_id,
                        roles=[UserRole.FRONTDESK, UserRole.NURSE],
                        topic=Topic.APPOINTMENT_RESCHEDULED,
                        priority=Priority.LOW,
                        title="Appointment updated",
                        body=f"{patient_name or 'A patient'} appointment updated to {new_when}.",
                        data=payload,
                        action_url=f"/facility/appointments/{appt.id}",
                        group_key=group_key,
                    )
        except Exception:
            pass

        return resp

    # ─────────────────────────────────────────────────────────────
    # Status transition actions
    # ─────────────────────────────────────────────────────────────

    @action(detail=True, methods=["POST"], url_path="check_in")
    def check_in(self, request, pk=None):
        if not has_facility_permission(request.user, 'can_check_in_appointments'):
            return Response(
                {"detail": "You do not have permission to check in appointments."},
                status=status.HTTP_403_FORBIDDEN
            )
        """
        Check in a patient for their appointment.
        
        Automatically creates a billing charge based on appointment type,
        but ONLY if facility has set a price for that appointment type.
        
        If no price is set:
        - Check-in still succeeds
        - No charge is created
        - Warning message returned
        """
        appointment = self.get_object()
        
        # Validate appointment can be checked in
        if appointment.status == "CHECKED_IN":
            return Response(
                {"detail": "Patient already checked in"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if appointment.status in ["CANCELLED", "NO_SHOW"]:
            return Response(
                {"detail": f"Cannot check in {appointment.status.lower()} appointment"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check in the patient
        appointment.status = "CHECKED_IN"
        appointment.save(update_fields=["status", "updated_at"])
        
        # Attempt to create billing charge
        charge_created = False
        charge_id = None
        charge_amount = None
        charge_error = None
        price_not_set = False
        
        try:
            # Get the service code based on appointment type
            service_code = f"APPT:{appointment.appt_type}"
            
            # Get the service
            service = Service.objects.filter(code=service_code).first()
            if not service:
                charge_error = f"Service not found for appointment type: {appointment.appt_type}"
            else:
                # Determine scope for price resolution
                facility = None
                owner = None
                
                if hasattr(appointment, 'facility') and appointment.facility:
                    facility = appointment.facility
                elif hasattr(appointment, 'provider') and appointment.provider:
                    # Get provider's facility or use as independent provider
                    if hasattr(appointment.provider, 'facility') and appointment.provider.facility:
                        facility = appointment.provider.facility
                    else:
                        owner = appointment.provider
                
                # Try to resolve the price
                try:
                    price_amount = resolve_price(
                        service=service,
                        facility=facility,
                        owner=owner,
                        system_hmo=appointment.patient.system_hmo if hasattr(appointment.patient, 'system_hmo') and appointment.patient.system_hmo else None,
                        tier=appointment.patient.hmo_tier if hasattr(appointment.patient, 'hmo_tier') and appointment.patient.hmo_tier else None,
                    )
                    
                    # If price is None or 0, it means no price has been set
                    if price_amount is None or price_amount == 0:
                        price_not_set = True
                        charge_error = f"No price configured for {appointment.get_appt_type_display()} appointments"
                    else:
                        # Format date from start_at (DateTimeField)
                        appt_date = appointment.start_at.strftime("%Y-%m-%d") if appointment.start_at else "Unknown date"
                        
                        # Create the charge
                        # NOTE: Charge model fields:
                        #   - patient, facility, owner, service (FKs)
                        #   - description, unit_price, qty, amount
                        #   - status, created_by, created_at
                        #   - NO 'currency' field (assumed NGN)
                        charge = Charge.objects.create(
                        patient=appointment.patient,
                        service=service,
                        unit_price=price_amount,  # ✅ Correct field
                        qty=1,                     # ✅ Required field
                        amount=price_amount,       # ✅ Total amount
                        description=f"{appointment.get_appt_type_display()} - {appt_date}",  # ✅ Correct date
                        facility=facility,
                        owner=owner,
                        status="UNPAID",           # ✅ Correct enum
                        created_by=request.user,   # ✅ Track creator
                    )
                        
                        charge_created = True
                        charge_id = charge.id
                        charge_amount = str(price_amount)
                        
                except Exception as price_error:
                    charge_error = f"Failed to resolve price: {str(price_error)}"
                    
        except Exception as e:
            charge_error = f"Billing error: {str(e)}"
        
        # Prepare response
        response_data = {
            "id": appointment.id,
            "status": appointment.status,
            "charge_created": charge_created,
        }
        
        if charge_created:
            response_data["charge_id"] = charge_id
            response_data["charge_amount"] = charge_amount
            response_data["message"] = "Patient checked in successfully. Billing charge created."
        elif price_not_set:
            response_data["charge_error"] = charge_error
            response_data["message"] = "Patient checked in. No charge created - please configure pricing for this appointment type."
            response_data["price_not_set"] = True
        else:
            response_data["charge_error"] = charge_error
            response_data["message"] = "Patient checked in. Billing charge could not be created automatically."
        
        return Response(response_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["GET"], url_path="service_prices")
    def service_prices(self, request):
        """
        Get pricing information for all appointment types.
        
        NO DEFAULT PRICES - Facilities must set their own prices.
        
        Returns a list of services with their pricing details:
        - service_id: Service database ID
        - service_code: Service code (e.g., "APPT:CONSULTATION")
        - service_name: Friendly service name
        - appt_type: Appointment type code
        - facility_price: Price set by facility (null if not set)
        - is_set: Whether facility has set a price
        
        Pricing is scoped to the current user's facility or provider profile.
        """
        user = request.user
        
        # Determine scope
        facility = None
        owner = None
        
        if hasattr(user, 'facility') and user.facility:
            facility = user.facility
        elif hasattr(user, 'provider_profile') and user.provider_profile:
            # Independent provider
            owner = user
        else:
            return Response(
                {"detail": "User has no facility or provider profile"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all appointment service codes
        service_codes = [
            "APPT:CONSULTATION",
            "APPT:FOLLOW_UP",
            "APPT:PROCEDURE",
            "APPT:DIAGNOSTIC_NON_LAB",
            "APPT:NURSING_CARE",
            "APPT:THERAPY_REHAB",
            "APPT:MENTAL_HEALTH",
            "APPT:IMMUNIZATION",
            "APPT:MATERNAL_CHILD_CARE",
            "APPT:SURGICAL_PRE_POST",
            "APPT:EMERGENCY_NON_ER",
            "APPT:TELEMEDICINE",
            "APPT:HOME_VISIT",
            "APPT:ADMIN_HMO_REVIEW",
            "APPT:LAB",
            "APPT:IMAGING",
            "APPT:PHARMACY",
            "APPT:OTHER",
        ]
        
        # Map appointment types to friendly names
        appt_type_names = {
            "CONSULTATION": "Consultation",
            "FOLLOW_UP": "Follow-up Visit",
            "PROCEDURE": "Procedure",
            "DIAGNOSTIC_NON_LAB": "Diagnostic (Non-Lab)",
            "NURSING_CARE": "Nursing Care",
            "THERAPY_REHAB": "Therapy/Rehabilitation",
            "MENTAL_HEALTH": "Mental Health",
            "IMMUNIZATION": "Immunization/Vaccination",
            "MATERNAL_CHILD_CARE": "Maternal/Child Care",
            "SURGICAL_PRE_POST": "Surgical (Pre/Post-op)",
            "EMERGENCY_NON_ER": "Emergency (Non-ER)",
            "TELEMEDICINE": "Telemedicine",
            "HOME_VISIT": "Home Visit",
            "ADMIN_HMO_REVIEW": "Administrative/HMO Review",
            "LAB": "Lab Visit",
            "IMAGING": "Imaging Visit",
            "PHARMACY": "Pharmacy Pickup",
            "OTHER": "Other",
        }
        
        results = []
        
        for service_code in service_codes:
            try:
                # Get or skip service
                service = Service.objects.filter(code=service_code).first()
                if not service:
                    continue
                
                # Extract appointment type from code
                appt_type = service_code.replace("APPT:", "")
                service_name = appt_type_names.get(appt_type, appt_type)
                
                # Get facility/owner specific price
                facility_price = None
                is_set = False
                
                # Check if facility/owner has set a custom price
                if facility:
                    price_obj = Price.objects.filter(
                        service=service,
                        facility=facility
                    ).first()
                    if price_obj:
                        facility_price = str(price_obj.amount)
                        is_set = True
                elif owner:
                    price_obj = Price.objects.filter(
                        service=service,
                        owner=owner
                    ).first()
                    if price_obj:
                        facility_price = str(price_obj.amount)
                        is_set = True
                
                results.append({
                    "service_id": service.id,
                    "service_code": service_code,
                    "service_name": service_name,
                    "appt_type": appt_type,
                    "facility_price": facility_price,  # null if not set
                    "is_set": is_set,  # whether price is set
                })
                
            except Exception as e:
                # Log error but continue with other services
                print(f"Error processing service {service_code}: {e}")
                continue
        
        return Response(results, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        """
        Cancel appointment.
        Patients can cancel their own; staff can cancel any in their facility.
        """
        appt = self.get_object()

        # Permission check
        if request.user.role != "PATIENT":
            self.permission_classes = [IsAuthenticated, IsStaff]
            self.check_permissions(request)
        else:
            # Patients can only cancel their own
            base_patient = getattr(request.user, "patient_profile", None)
            if not base_patient:
                return Response(
                    {"detail": "No patient profile linked."}, status=403
                )
            allowed_ids = {base_patient.id} | set(
                base_patient.dependents.values_list("id", flat=True)
            )
            if appt.patient_id not in allowed_ids:
                return Response(
                    {"detail": "You can only cancel your own appointments."},
                    status=403,
                )

        if appt.status == ApptStatus.COMPLETED:
            return Response(
                {"detail": "Completed appointments cannot be cancelled."},
                status=400,
            )

        appt.status = ApptStatus.CANCELLED
        appt.save(update_fields=["status", "updated_at"])

        try:
            when = appt.start_at.strftime("%Y-%m-%d %H:%M") if appt.start_at else ""
            payload = {"appointment_id": appt.id, "patient_id": appt.patient_id}
            patient_name = " ".join([p for p in [getattr(appt.patient, 'first_name', ''), getattr(appt.patient, 'middle_name', ''), getattr(appt.patient, 'last_name', '')] if p]).strip()
            group_key = f"APPT:{appt.id}:CANCELLED"

            if appt.provider_id:
                notify_user(
                    user=appt.provider,
                    topic=Topic.APPOINTMENT_CANCELLED,
                    priority=Priority.NORMAL,
                    title="Appointment cancelled",
                    body=f"An appointment ({when}) was cancelled.",
                    facility_id=appt.facility_id,
                    data=payload,
                    action_url="/facility/appointments",
                    group_key=group_key,
                )

            if appt.patient:
                notify_patient(
                    patient=appt.patient,
                    topic=Topic.APPOINTMENT_CANCELLED,
                    priority=Priority.LOW,
                    title="Appointment cancelled",
                    body="Your appointment was cancelled.",
                    facility_id=appt.facility_id,
                    data=payload,
                    action_url="/patient/appointments",
                    group_key=group_key,
                )
            # Ops feed (role-scoped): Frontdesk gets cancellation events
            if appt.facility_id:
                cancelled_by = "patient" if request.user.role == "PATIENT" else "staff"
                notify_facility_roles(
                    facility_id=appt.facility_id,
                    roles=[UserRole.FRONTDESK],
                    topic=Topic.APPOINTMENT_CANCELLED,
                    priority=Priority.NORMAL,
                    title="Appointment cancelled",
                    body=f"{patient_name}'s appointment ({when}) was cancelled by {cancelled_by}.",
                    data=payload,
                    action_url=f"/facility/appointments/{appt.id}",
                    group_key=group_key,
                )

        except Exception:
            pass

        return Response(
            AppointmentSerializer(appt, context={"request": request}).data
        )

    @action(detail=True, methods=["post"])
    def no_show(self, request, pk=None):
        """Mark appointment as no-show (patient didn't arrive)."""
        appt = self.get_object()
        self.permission_classes = [IsAuthenticated, IsStaff]
        self.check_permissions(request)

        if appt.status != ApptStatus.SCHEDULED:
            return Response(
                {"detail": "Only scheduled appointments can be marked no-show."},
                status=400,
            )

        appt.status = ApptStatus.NO_SHOW
        appt.save(update_fields=["status", "updated_at"])

        try:
            when = appt.start_at.strftime("%Y-%m-%d %H:%M") if appt.start_at else ""
            payload = {"appointment_id": appt.id, "patient_id": appt.patient_id}
            group_key = f"APPT:{appt.id}:NO_SHOW"
            if appt.provider_id:
                notify_user(
                    user=appt.provider,
                    topic=Topic.APPOINTMENT_NO_SHOW,
                    priority=Priority.NORMAL,
                    title="Appointment no-show",
                    body=f"Patient did not show for {when}.",
                    facility_id=appt.facility_id,
                    data=payload,
                    action_url="/facility/appointments",
                    group_key=group_key,
                )
            if appt.patient:
                notify_patient(
                    patient=appt.patient,
                    topic=Topic.APPOINTMENT_NO_SHOW,
                    priority=Priority.LOW,
                    title="Missed appointment",
                    body="You missed your scheduled appointment.",
                    facility_id=appt.facility_id,
                    data=payload,
                    action_url="/patient/appointments",
                    group_key=group_key,
                )
        except Exception:
            pass

        # Patient email (opt-in)
        try:
            send_no_show(appt)
        except Exception:
            pass

        return Response(
            AppointmentSerializer(appt, context={"request": request}).data
        )

    # ─────────────────────────────────────────────────────────────
    # Utility endpoints
    # ─────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"])
    def statuses(self, request):
        """Return available appointment statuses."""
        return Response([{"value": c, "label": l} for c, l in ApptStatus.choices])

    @action(detail=False, methods=["get"])
    def summary(self, request):
        """
        Return appointment counts by status for the current user's scope.
        Useful for dashboard widgets.
        """
        qs = self.get_queryset()
        
        # Optional date range
        date_filter = request.query_params.get("date", "today")
        today = timezone.now().date()
        
        if date_filter == "today":
            qs = qs.filter(start_at__date=today)
        elif date_filter == "this_week":
            week_start = today - timezone.timedelta(days=today.weekday())
            week_end = week_start + timezone.timedelta(days=6)
            qs = qs.filter(start_at__date__gte=week_start, start_at__date__lte=week_end)
        
        counts = {
            "total": qs.count(),
            "scheduled": qs.filter(status=ApptStatus.SCHEDULED).count(),
            "checked_in": qs.filter(status=ApptStatus.CHECKED_IN).count(),
            "completed": qs.filter(status=ApptStatus.COMPLETED).count(),
            "cancelled": qs.filter(status=ApptStatus.CANCELLED).count(),
            "no_show": qs.filter(status=ApptStatus.NO_SHOW).count(),
        }
        
        return Response(counts)

    @action(detail=False, methods=["post"])
    def send_reminders(self, request):
        """
        Send reminder notifications for appointments in a time range.
        Used by scheduled jobs.
        """
        start = parse_datetime(request.data.get("start")) or request.data.get("start")
        end = parse_datetime(request.data.get("end")) or request.data.get("end")

        if not start or not end:
            return Response({"detail": "start and end required"}, status=400)

        qs = self.get_queryset().filter(
            status=ApptStatus.SCHEDULED, start_at__gte=start, start_at__lte=end
        )

        n = 0
        for appt in qs:
            send_reminder(appt)
            if appt.patient:
                notify_patient(
                    patient=appt.patient,
                    topic="APPT_REMINDER",
                    title="Appointment Reminder",
                    body=f"Reminder: appointment at {appt.start_at}.",
                    data={"appointment_id": appt.id},
                    facility_id=appt.facility_id,
                    allow_email=getattr(appt, "notify_email", False),
                )
            n += 1

        return Response({"sent": n})

    @action(detail=False, methods=["get"], url_path="hmo-catalog")
    def hmo_catalog(self, request):
        """
        Return appointment services with HMO-specific pricing.
        
        Query params:
        - hmo_id (required): SystemHMO ID to get pricing for
        - tier_id (optional): HMOTier ID for tier-specific pricing
        
        Returns list of appointment services with:
        - catalog_price (facility's default price)
        - hmo_price (HMO-specific override or falls back to catalog)
        - tier information if tier_id is provided
        - pricing resolution: tier-specific → HMO-default → facility-default
        
        ✅ FIXED: Now returns exactly 18 appointment services
        """
        from billing.models import HMOPrice, Service, Price
        from patients.models import SystemHMO, HMOTier

        hmo_id = request.query_params.get("hmo_id")
        tier_id = request.query_params.get("tier_id")
        
        if not hmo_id:
            return Response({"detail": "hmo_id is required"}, status=400)

        # Get SystemHMO (system-wide, no facility filter)
        try:
            system_hmo = SystemHMO.objects.get(id=hmo_id, is_active=True)
        except SystemHMO.DoesNotExist:
            return Response({"detail": "HMO not found"}, status=404)
        
        # Get tier if specified
        tier = None
        if tier_id:
            try:
                tier = HMOTier.objects.get(id=tier_id, system_hmo=system_hmo, is_active=True)
            except HMOTier.DoesNotExist:
                return Response({"detail": "Tier not found or doesn't belong to this HMO"}, status=404)

        # Get facility if user is facility-scoped
        facility_id = getattr(request.user, "facility_id", None)

        # Appointment services can exist in two formats depending on when the facility was set up:
        # - New system: codes prefixed with "APPT:" (seed_appointment_services)
        # - Legacy system: consultation/service codes without the prefix (e.g., CONSULT_STD)
        # We include both so facilities see their full appointment catalog in the HMO pricing tab.
        LEGACY_APPT_CODES = [
            "CONSULT_STD",
            "CONSULT_FOLLOW_UP",
            "CONSULT_EMERGENCY",
            "CONSULT_SPECIALIST",
            "CONSULT_PEDIATRIC",
            "ANNUAL_CHECKUP",
            "PHYSICAL_EXAM",
            "WELLNESS_VISIT",
            "LAB_COLLECTION",
            "X_RAY_SCREENING",
            "DENTAL_CHECKUP",
            "VISION_SCREENING",
            "HEARING_TEST",
            "COUNSELING_SESSION",
            "NUTRITION_CONSULT",
            "THERAPY_SESSION",
        ]

        services = Service.objects.filter(
            Q(code__startswith="APPT:") | Q(code__in=LEGACY_APPT_CODES),
            is_active=True,
        ).order_by("code")

        # ✅ FIX #2: Updated service name mapping for original service codes
        # The old mapping only had alternative service names (CONSULT_STD, etc.)
        # This now includes the original service codes (CONSULTATION, FOLLOW_UP, etc.)
        service_names = {
            # Original service codes from seed_appointment_services.py
            "CONSULTATION": "Consultation",
            "FOLLOW_UP": "Follow-up Visit",
            "PROCEDURE": "Procedure",
            "DIAGNOSTIC_NON_LAB": "Diagnostic (Non-Lab)",
            "NURSING_CARE": "Nursing Care",
            "THERAPY_REHAB": "Therapy/Rehabilitation",
            "MENTAL_HEALTH": "Mental Health",
            "IMMUNIZATION": "Immunization/Vaccination",
            "MATERNAL_CHILD_CARE": "Maternal/Child Care",
            "SURGICAL_PRE_POST": "Surgical (Pre/Post-op)",
            "EMERGENCY_NON_ER": "Emergency (Non-ER)",
            "TELEMEDICINE": "Telemedicine",
            "HOME_VISIT": "Home Visit",
            "ADMIN_HMO_REVIEW": "Administrative/HMO Review",
            "LAB": "Lab Visit",
            "IMAGING": "Imaging Visit",
            "PHARMACY": "Pharmacy Pickup",
            "OTHER": "Other",
            
            # Alternative service codes (kept for backward compatibility)
            "CONSULT_STD": "Standard Consultation",
            "CONSULT_FOLLOW_UP": "Follow-up Consultation",
            "CONSULT_EMERGENCY": "Emergency Consultation",
            "CONSULT_SPECIALIST": "Specialist Consultation",
            "CONSULT_PEDIATRIC": "Pediatric Consultation",
            "ANNUAL_CHECKUP": "Annual Health Checkup",
            "PHYSICAL_EXAM": "Physical Examination",
            "WELLNESS_VISIT": "Wellness Visit",
            "LAB_COLLECTION": "Lab Sample Collection",
            "X_RAY_SCREENING": "X-Ray Screening",
            "DENTAL_CHECKUP": "Dental Checkup",
            "VISION_SCREENING": "Vision Screening",
            "HEARING_TEST": "Hearing Test",
            "COUNSELING_SESSION": "Counseling Session",
            "NUTRITION_CONSULT": "Nutrition Consultation",
            "THERAPY_SESSION": "Therapy Session (Physical/Occupational)",
        }

        result = []
        for service in services:
            code = service.code
            service_type = code.replace("APPT:", "")

            # Get facility price if applicable
            catalog_price = service.default_price
            if facility_id:
                facility_price = Price.objects.filter(
                    facility_id=facility_id, service=service
                ).first()
                if facility_price:
                    catalog_price = facility_price.amount

            # Get HMO price - try tier-specific first, fall back to HMO-level
            hmo_price_obj = None
            has_tier_specific = False
            
            if tier and facility_id:
                # Try tier-specific price first
                hmo_price_obj = HMOPrice.objects.filter(
                    facility_id=facility_id,
                    system_hmo=system_hmo,
                    tier=tier,
                    service=service,
                    is_active=True
                ).first()
                
                if hmo_price_obj:
                    has_tier_specific = True
            
            if not hmo_price_obj and facility_id:
                # Fall back to HMO-level default (no tier)
                hmo_price_obj = HMOPrice.objects.filter(
                    facility_id=facility_id,
                    system_hmo=system_hmo,
                    tier__isnull=True,
                    service=service,
                    is_active=True
                ).first()

            hmo_price = hmo_price_obj.amount if hmo_price_obj else catalog_price

            # Calculate discount if HMO price is different
            discount = 0
            if hmo_price < catalog_price:
                discount = ((catalog_price - hmo_price) / catalog_price) * 100

            # ✅ UNCHANGED: Return format remains the same, frontend compatibility maintained
            result.append({
                "service_id": service.id,
                "service_code": code,
                "service_type": service_type,
                "service_name": service_names.get(service_type, service.name),
                "catalog_price": str(catalog_price),
                "duration": None,
                
                # HMO pricing
                "hmo_id": system_hmo.id,
                "hmo_name": system_hmo.name,
                "tier_id": tier.id if tier else None,
                "tier_name": tier.name if tier else None,
                "hmo_price": str(hmo_price),
                "has_tier_specific_price": has_tier_specific,
                "discount": round(discount, 2),
            })

        return Response(result)

    @action(detail=False, methods=["post"], url_path="set-hmo-price", permission_classes=[IsAuthenticated, IsStaff])
    def set_hmo_price(self, request):
        """
        Set HMO-specific price for an appointment service.
        
        Body:
        {
            "hmo_id": 1,           # SystemHMO ID (required)
            "tier_id": 2,          # HMOTier ID (optional) - if provided, sets tier-specific price
            "service_id": 5,       # or service_code
            "amount": "2000.00"
        }
        
        Pricing levels:
        - No tier_id: Sets HMO-level default (applies to all tiers unless overridden)
        - With tier_id: Sets tier-specific price (only for that tier)
        """
        from billing.models import HMOPrice

        # Only admins can set HMO prices
        if request.user.role not in (UserRole.SUPER_ADMIN, UserRole.ADMIN):
            return Response({"detail": "Only admins can set HMO prices"}, status=403)

        hmo_id = request.data.get("hmo_id")
        tier_id = request.data.get("tier_id")
        service_id = request.data.get("service_id")
        service_code = request.data.get("service_code")
        amount = request.data.get("amount")

        if not all([hmo_id, (service_id or service_code), amount]):
            return Response(
                {"detail": "hmo_id, (service_id or service_code), and amount are required"},
                status=400,
            )

        # Get SystemHMO (system-wide, no facility filter)
        try:
            system_hmo = SystemHMO.objects.get(id=hmo_id, is_active=True)
        except SystemHMO.DoesNotExist:
            return Response({"detail": "HMO not found"}, status=404)
        
        # Get tier if specified
        tier = None
        if tier_id:
            try:
                tier = HMOTier.objects.get(id=tier_id, system_hmo=system_hmo, is_active=True)
            except HMOTier.DoesNotExist:
                return Response({"detail": "Tier not found or doesn't belong to this HMO"}, status=404)

        # Get service
        if service_id:
            try:
                service = Service.objects.get(id=service_id)
            except Service.DoesNotExist:
                return Response({"detail": "Service not found"}, status=404)
        else:
            try:
                service = Service.objects.get(code=service_code)
            except Service.DoesNotExist:
                return Response({"detail": "Service not found"}, status=404)

        # Get facility - required for HMO pricing
        facility_id = getattr(request.user, "facility_id", None)
        if not facility_id:
            return Response(
                {"detail": "Facility is required for HMO pricing"},
                status=400,
            )

        # Create or update HMO price with tier support
        try:
            hp, created = HMOPrice.objects.update_or_create(
                facility_id=facility_id,
                system_hmo=system_hmo,
                tier=tier,  # Can be None for HMO-level default
                service=service,
                defaults={
                    "amount": amount,
                    "is_active": True
                }
            )

            return Response(
                {
                    "success": True,
                    "service_id": service.id,
                    "service_code": service.code,
                    "service_name": service.name,
                    "hmo_id": system_hmo.id,
                    "hmo_name": system_hmo.name,
                    "tier_id": tier.id if tier else None,
                    "tier_name": tier.name if tier else None,
                    "hmo_price": str(hp.amount),
                    "is_tier_specific": bool(tier),
                    "created": created,
                    "message": f"{'Created' if created else 'Updated'} {'tier-specific' if tier else 'HMO-level default'} price"
                }
            )
        except Exception as e:
            return Response({"detail": str(e)}, status=500)


    @action(detail=False, methods=["post"], url_path="import-hmo-file", permission_classes=[IsAuthenticated, IsStaff])
    def import_hmo_file(self, request):
        """
        Bulk import HMO prices for appointment services from CSV/Excel.

        Query params:
        - hmo_id (required): SystemHMO ID
        - tier_id (optional): HMOTier ID (tier-specific pricing)

        File columns (required):
        - code: Service code (e.g. APPT:CONSULTATION or CONSULT_STD)
        - price: HMO price

        Optional columns:
        - name: Service name (ignored if present)

        Returns:
        - created, updated, errors
        """
        import csv
        import io
        from decimal import Decimal, InvalidOperation

        from billing.models import HMOPrice, Service
        from patients.models import SystemHMO, HMOTier

        # Facility only
        facility_id = getattr(request.user, "facility_id", None)
        if not facility_id:
            return Response({"detail": "Facility is required for HMO pricing"}, status=400)

        hmo_id = request.query_params.get("hmo_id") or request.query_params.get("hmo")
        tier_id = request.query_params.get("tier_id")

        if not hmo_id:
            return Response({"detail": "hmo_id query parameter is required"}, status=400)

        system_hmo = SystemHMO.objects.filter(id=hmo_id, is_active=True).first()
        if not system_hmo:
            return Response({"detail": "HMO not found"}, status=404)

        tier = None
        if tier_id:
            tier = HMOTier.objects.filter(id=tier_id, system_hmo=system_hmo, is_active=True).first()
            if not tier:
                return Response({"detail": "Tier not found or doesn't belong to this HMO"}, status=404)

        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "file is required"}, status=400)

        # Parse file (CSV or Excel)
        try:
            file_ext = file_obj.name.lower().split(".")[-1]

            if file_ext == "csv":
                content = file_obj.read().decode("utf-8", errors="ignore")
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)

            elif file_ext in ("xlsx", "xls"):
                if file_ext == "xls":
                    return Response({"detail": "Legacy .xls is not supported. Please save as .xlsx or CSV."}, status=400)

                import openpyxl
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                ws = wb.active

                headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: row[i] for i in range(len(headers))})

            else:
                return Response({"detail": "File must be CSV or Excel (.xlsx)"}, status=400)

        except Exception as e:
            return Response({"detail": f"Failed to parse file: {str(e)}"}, status=400)

        created_count = 0
        updated_count = 0
        errors = []

        for idx, row in enumerate(rows, start=2):
            try:
                # Read fields (accept a few aliases)
                code_raw = (row.get("code") or row.get("service_code") or row.get("service") or "")
                code_raw = str(code_raw).strip().upper()
                if not code_raw:
                    errors.append(f"Row {idx}: Missing code")
                    continue

                price_raw = (row.get("price") or row.get("amount") or "")
                price_raw = str(price_raw).strip()
                if not price_raw:
                    errors.append(f"Row {idx}: Missing price")
                    continue

                try:
                    amount = Decimal(price_raw)
                    if amount < 0:
                        errors.append(f"Row {idx}: Price cannot be negative")
                        continue
                except (InvalidOperation, ValueError):
                    errors.append(f"Row {idx}: Invalid price '{price_raw}'")
                    continue

                # Find service (try exact, then APPT: prefix fallback)
                service = Service.objects.filter(code=code_raw).first()
                if not service and not code_raw.startswith("APPT:"):
                    service = Service.objects.filter(code=f"APPT:{code_raw}").first()

                if not service and code_raw.startswith("APPT:"):
                    service = Service.objects.filter(code=code_raw.replace("APPT:", "", 1)).first()

                if not service:
                    errors.append(f"Row {idx}: Service '{code_raw}' not found")
                    continue

                hp, created = HMOPrice.objects.update_or_create(
                    facility_id=facility_id,
                    system_hmo=system_hmo,
                    tier=tier,  # None means HMO-level default
                    service=service,
                    defaults={"amount": amount, "is_active": True},
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")

        return Response(
            {
                "created": created_count,
                "updated": updated_count,
                "errors": errors,
                "message": f"Processed {created_count + updated_count} prices",
            },
            status=200,
        )
# ─────────────────────────────────────────────────────────────
# Utility functions for syncing appointment status from encounters
# ─────────────────────────────────────────────────────────────

def sync_appointment_on_encounter_start(appointment_id: int):
    """
    Called when an encounter is started from an appointment.
    Updates appointment status to CHECKED_IN if still SCHEDULED.
    """
    try:
        appt = Appointment.objects.get(id=appointment_id)
        if appt.status == ApptStatus.SCHEDULED:
            appt.status = ApptStatus.CHECKED_IN
            appt.save(update_fields=["status", "updated_at"])
    except Appointment.DoesNotExist:
        pass


def sync_appointment_on_encounter_close(encounter_id: int):
    """
    Called when an encounter is closed.
    Updates linked appointment status to COMPLETED if not already terminal.
    """
    try:
        appt = Appointment.objects.filter(encounter_id=encounter_id).first()
        if appt and appt.status not in (
            ApptStatus.COMPLETED,
            ApptStatus.CANCELLED,
            ApptStatus.NO_SHOW,
        ):
            appt.status = ApptStatus.COMPLETED
            appt.save(update_fields=["status", "updated_at"])
    except Exception:
        pass
